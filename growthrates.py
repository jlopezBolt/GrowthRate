import untangle, argparse
import sys
from argparse import ArgumentParser
from collections import defaultdict
import numpy
from openpyxl import Workbook
from openpyxl.charts import LineChart, Reference, Series, ErrorBar, ScatterChart
import xlsxwriter
import re
import os.path
import collections
import xml.etree.ElementTree as ET
import matplotlib.pyplot as plt
import re
import pylab
from mpl_toolkits.mplot3d import Axes3D
import scipy.stats as t
import numpy as np
from sklearn import linear_model
import math
from math import log
import csv
import itertools
from gooey import Gooey, GooeyParser
from itertools import groupby
import datetime
from itertools import takewhile
import xlrd
from xlutils.copy import copy
from xlsx2csv import *


########### Variables  ########

input = []
input_array = []
time_points = []
strain_name_list = []
blank_wells = []
dilution = []

###################### Functions #######################################################################

def plate_averages(well_value, plate_coordinate,input):
	value_ave = [well_value[x:x+96] for x in xrange(0, len(well_value), 96)]
	well_ave = [plate_coordinate[x:x+96] for x in xrange(0, len(plate_coordinate), 96)]

	num_plates = len(value_ave)
	num_strains = len(input[2])
	strain_counter = 1
	time_points = input[0]
	averages_out = []
	stdv_out = []
	while strain_counter < num_strains:
		averages = []
		blank_wells = input[3]
	
		plate_counter = 0
		current_strain = input[2][strain_counter]
		current_strain_wells = input[strain_counter+3]
		strain_averages_out = []
		strain_stdv_out = []
		corrected_strain_values =[]
		while plate_counter < num_plates:
			current_time = time_points[plate_counter]
			current_plate_dil = float(input[1][plate_counter])
		
			current_plate_data = map(None, well_ave[plate_counter], value_ave[plate_counter])
		
			plate_blank_values = []
			for x in current_plate_data:
				if x[0] in blank_wells:
					plate_blank_values.append(float(x[1]))
			average_blank_value = average(plate_blank_values)

			counter = 0
			plate_strain_values = []
			for x in current_plate_data:
				if x[0] in current_strain_wells:
					plate_strain_values.append(float(x[1]))
					
			corrected_strain_values_temp = []
			for x in plate_strain_values:
				corrected_strain_values_temp.append(float((x-average_blank_value)*current_plate_dil))

			corrected_strain_values.append(corrected_strain_values_temp)
	
			current_strain_average = average(corrected_strain_values_temp)
			current_strain_stdv = np.std(corrected_strain_values_temp)
			
			current_strain_out = (current_strain, current_time, current_strain_average, current_strain_stdv)
			
			strain_averages_out.append(current_strain_out)
			
			plate_counter = plate_counter+1
			
		averages_out.append(strain_averages_out)
		
		strain_counter = strain_counter+1
	return averages_out



'''
Returns the measurement value, time stamp, and well for given wells from data matrix

inputeters
-----------------
data : 
'''
def strain_data_extract(data, strain_pos):
	s_value = []
	s_time = []
	s_cell = []
	for sub_list in data:
		if sub_list[1] in strain_pos:
			s_value.append(float(sub_list[2])) # makes a list of all the values found in wells matching above criteria
			s_time.append(sub_list[0])
			s_cell.append(sub_list[1])
			
	return s_value, s_time, s_cell

		
		

	
##########################################################################
'''
Removes everthing that is not followed by 'hr'

'''
def time_cleanup(time_point):
	timelist = []
	for sublist in time_point:
		hr = re.compile('..hr')
		h = hr.search(sublist)
		time = str(h.group())
		timelist.append(time)
		
	count = 0
	temp = []
	hour_numbers = []

	while count < len(timelist):
		temp= float(re.search(r'\d+', timelist[count]).group())
		count = count + 1
		hour_numbers.append(temp)
	return hour_numbers
###############################################################################

def average(values): 
	average = []
	average = sum(values)/len(values)
	return average
	
###############################################################################
'''
Calculates the predicted y and then
the deviations from predicted from observed
y(observed) - y(predicted)
Returns the diference of observed and predicted values of y (OD)

'''
def y_err1(time, slope, intercept, values):
	y_error= []
	temp = [x * slope for x in time]
	y_error = [b + intercept for b in temp]
	y_error = np.subtract(values, y_error)
	return y_error
	
###############################################################################

def conf_calc(x, y_err, c_limit=0.975):
    '''
    Calculates confidence interval of regression between x and y
    
    inputeters
    ----------
    x:       1D numpy array
    y_err:   1D numpy array of residuals (y - fit)
    c_limit: (optional) float number representing the area to the left
             of the critical value in the t-statistic table
             eg: for a 2 tailed 95% confidence interval (the default)
                    c_limit = 0.975
    Returns
    -------
    confs: 1D numpy array of predicted y values for x inputs
    
    '''
    # Define the variables you need
    # to calculate the confidence interval
    mean_x = np.mean(x)			# mean of x
    n = len(x)				# number of samples in origional fit
    tstat = t.t.ppf(c_limit, n-1)         # appropriate t value
    s_err = np.sum(np.power(y_err,2))	# sum of the squares of the residuals

    # create series of new test x-values to predict for
    p_x = np.linspace(np.min(x),np.max(x),100)

    confs = tstat * np.sqrt((s_err/(n-2))*(1.0/n + (np.power((p_x-mean_x),2)/
			((np.sum(np.power(x,2)))-n*(np.power(mean_x,2))))))

    return p_x, confs

###############################################################################  
'''
estimates a linear model using an increasing number of time points. It starts with a minimum of three time points. 
It returns the timepoint that gives the highest r value for the linear regression.

'''

def time_select(hour, log_od_value, input_array):
	time_min = (len(input_array[3])-1)*3
	time_max = len(hour)
	r_values = []
	slopes = []
	intercepts = []
	p_values = []
	std_errs = []
	slope_r_value = []
	time = []
	hour_max = []
	time_start = len(input_array[3])-1
	time_start_list = []
	while time_min <= time_max:
		slope, intercept, r_value, p_value, std_err = t.linregress(hour[time_start:time_min], log_od_value[time_start:time_min])
		r_values.append(r_value)
		slopes.append(slope)
		intercepts.append(intercept)
		p_values.append(p_value)
		std_errs.append(std_err)
		hour_max.append(hour[time_min])
		time.append(time_min)
		time_start_list.append(time_start)
		

		slope_r_value = zip(slopes, r_values, time, std_errs, hour_max, time_start_list)
		
		time_min = time_min + len(input_array[3])
		
		if time_start == 0:
			time_start = time_start + len(input_array)-1
		else:	
			time_start = time_start + len(input_array)
		
	for x in slope_r_value:
		if max(slopes) == x[0]:
			max_info = x
			hour_max_out = x[4]
			time_start_out = x[5]
			
	return max_info, hour_max_out, time_start_out 





###############################################################################  



def ylines_calc(p_x, confs, slope, intercept):
    '''
    Calculates the three lines that will be plotted
    
    inputeters
    ----------
    p_x:   1D array with values spread evenly between min(x) and max(x)
    confs: 1D array with confidence values for each value of p_x
    slope
    intercept   
    
    Returns
    -------
    p_y:    1D array with values corresponding to fit line (for p_x values)
    upper:  1D array, values corresponding to upper confidence limit line
    lower:  1D array, values corresponding to lower confidence limit line
    
    '''
    # now predict y based on test x-values
    p_y = [x * slope for x in p_x]
    p_y = [b + intercept for b in p_y]
    
    
    # get lower and upper confidence limits based on predicted y and confidence intervals
    lower = p_y - abs(confs)
    upper = p_y + abs(confs)

    return p_y, lower, upper
    
###############################################################################  

def plot_linreg_CIs(x, y, p_x, p_y, lower, upper, strain_name):
	minimum = min(y) + min(y)*.10
	maximum = max(y) + max(y)*.10
	plt.scatter(x,y)
	plt.xlabel('Hours')
	plt.ylabel('Log(OD)')
	plt.ylim((minimum, maximum))
	plt.title('growth curve and confidence limits for RMs%s' % strain_name)
	plt.plot(x,y,'bo',label='Sample observations')
	# plot line of best fit
	plt.plot(p_x,p_y,'r-',label='Regression line')
	# plot confidence limits
	plt.plot(p_x,lower,'b--',label='Lower confidence limit (95%)')
	plt.plot(p_x,upper,'b--',label='Upper confidence limit (95%)')
	
	plt.legend(bbox_to_anchor=(1, .2), prop={'size':10})
	
	plt.savefig('Growth Rate for %s' % strain_name)
	#plt.show()
	plt.close()

	

###############################################################################  

@Gooey(program_name = "growthrates3")
def parse_args():
	parser = GooeyParser(description="Growth Curve Analysis")
	parser.add_argument('-f', '--elisa_file', help="xml file from plate reader", type=str, default=None, required=True, metavar="File", widget='FileChooser')
	parser.add_argument('-o', '--out_file', help="name of output xlsx file", type=str, default=None, required=True, widget='DirChooser')
	parser.add_argument('-i', '--data_info', help = "file with data information", type=str, default = None, required = True, widget = 'FileChooser')
	#parser.add_argument('-c', '--graphs', help="name of output xlsx file for graphs", type=str, default=None, required=True)
	parser.add_argument('-d', '--out_directory', help="directory for outfiles: platemap & correctedOD", type=str, default=None, required = True, widget = 'DirChooser')

	args = parser.parse_args()

	if os.path.isfile(args.elisa_file) == True:
		try:
			data_obj = untangle.parse(args.elisa_file)
		except Exception:
			print("The file %s is not readable!" % args.elisa_file)
			sys.exit(0)
		plate_name = data_obj.Experiment.PlateSections[1].PlateSection['Name']
		return args

args = parse_args()

tree = ET.parse(args.elisa_file)
root = tree.getroot()
out_dir=args.out_directory
out_file = args.out_file
plate_names = {}

plate_names = defaultdict(list)





############ Read Plate Layeout, Strains, dilution factor, and time points ################

reader_input_template = csv.reader(open(args.data_info, 'rU'))
for row in reader_input_template:
    input.append(row)

for sublist in input:
    sublist = filter(None, sublist)
    input_array.append(sublist)

   
for sublist in input_array[0]:
    time_points.append(sublist)

all_wells = []
counter_all = 3
while counter_all < len(input_array):														## creates a nested list with each sublist containing the name of the wells for each strain and controls
	all_wells.append(input_array[counter_all])
	counter_all = counter_all+1

for sublist in input_array[2]:       
    strain_name_list.append(sublist)

strain_list = []
for sublist in input_array[2]:
	strain_list.append(sublist)

for sublist in input_array[3]:
    blank_wells.append(sublist)

counter = 4
strain_wells = []

while (counter < (len(input_array))):
    strain_wells.append((input_array[counter]))
    counter = counter+1

for sublist in input_array[1]:
	dilution.append(sublist)

dilution_factor =[]
temp1 = []
dilution = list(map(float, dilution))
for sublist in dilution:
	temp1 = [sublist] * 96
	dilution_factor.extend(temp1)


hour = []
temp =[]
time_points = list(map(float, time_points))
for sublist in time_points:
    temp = [sublist] * len(input_array[3])
    hour.extend(temp)

######################################################################

#extracts data values from plates
well_value =[]
well_name = []
plate_name = []
read_time = []
raw_times = []
	
for plate in root.getiterator('PlateSection'):
	raw_times.append(plate.attrib.get('ReadTime'))
	for info in plate.getiterator('Wavelengths'):
		for well in info:
			for node in well:
				for data in node:
					for value in data:	
						read_time.append(plate.attrib.get('ReadTime'))
						plate_name.append(plate.attrib.get('Name'))
						well_value.append(value.text)
						well_name.append(data.get('Name'))

 

well_time = []
for x in read_time:																			## creates a list that has the read time for each well
	well_time_temp = datetime.datetime.strptime(x,'%I:%M %p %m/%d/%Y')
	well_time.append(well_time_temp)

												
matrix = [plate_name, well_name, well_value, dilution_factor]
data = zip(*matrix)




																							## the name of the well, the value in that well, and the dilution factor for that well
###########
platemap_location = out_dir + "/platemap.xlsx"
workbook = xlsxwriter.Workbook(platemap_location)
ws = workbook.add_worksheet()
ws.write(0,0, 'Well')
ws.write(0,1, 'Strain')
ws.write(0,2, 'Environment')
row_counter = 1

counter = 0
strain_list_long = []
while counter<len(all_wells):
	mult = len(all_wells[counter])
	temp_strain = [strain_list[counter]]*mult
	strain_list_long.append(temp_strain)
	counter = counter+1

row_counter = 1
strain_counter = 0
while strain_counter < len(strain_list_long):
	ws.write_column(row_counter, 0, all_wells[strain_counter])
	ws.write_column(row_counter, 1, strain_list_long[strain_counter])
	row_counter = row_counter + len(all_wells[strain_counter])
	strain_counter= strain_counter +1
	
workbook.close()



################ Analysis #######################################################################
counter = 0
slope_list = []
r_value_list = []
std_err_list = []
od_value_out = []
log_od_value_out = []
time_used = []
hour_max_list = []
average_matrix =[]
A600_value_out = []
time_start = 0

blank_value, blank_time, blank_well = strain_data_extract(data, blank_wells)				## extracts the info for - controls
blank_average = average(blank_value)


for wells in strain_wells:																	# wells is a list of well coordinates for given strain
	od_value = []
	A600_value = []
	log_od_value = []
	for sublist in data:															# sublist in data is list [time stamp, well, value, dilution factor
		if sublist[1] in wells:
			A600_value.append(float(sublist[2]))
			#print sublist[1]															# iterates over individual wells found in well coordinate list of strain
			od_value.append((float(sublist[2])-blank_average)*(sublist[3]))					# calculates the corrected value and adds it to od_value
			log_od_value = [np.log(y) for y in od_value]									# Takes the log of the corrected OD value 
	A600_value_out.append(A600_value)
	od_value_out.append(od_value)
	

	
	log_od_value_out.append(log_od_value)
	max_info, hour_max, time_start  = time_select(hour, log_od_value, input_array)

	time_constraint = max_info[2]
	
	time_used.append(max_info[2])
	
	time_1 = max_info[5]
	hour_max_list.append(hour_max)
	
	hour_set = []
	for element in hour:
		if element not in hour_set:
			hour_set.append(element)
 		
	slope, intercept, r_value, p_value, std_err = t.linregress(hour[time_1:time_constraint], log_od_value[time_1:time_constraint])				# Linear regression using timepoints in hour and the log OD values. Returns regression estimates and inputeters
	

	y_error = y_err1(hour[time_start:time_constraint], slope, intercept,log_od_value[time_start:time_constraint])											# Calculates error in y
	p_x, confs = conf_calc(hour[time_start:time_constraint], y_error, 0.975)																		# Calculates Confidence values for different values of x
	p_y, lower, upper = ylines_calc(p_x, confs, slope, intercept)	
	plot_linreg_CIs(hour[time_start:time_constraint], log_od_value[time_start:time_constraint], p_x, p_y, lower, upper, strain_name_list[counter+1])
	slope_list.append(slope)
	r_value_list.append(r_value)
	std_err_list.append(std_err)
	counter = counter + 1
	

wells_in_plate = well_name[0:96]															# Creates a list that contains the names for 1 plate

output = zip(strain_name_list[1:], slope_list, r_value_list, std_err_list)

dates = []
for x in raw_times:																			## reformats the raw_times into a python date object
	date = datetime.datetime.strptime(x,'%I:%M %p %m/%d/%Y')								
	dates.append(date)
#print dates

delta_time_points = []
time_counter = 1

### calculates the time difference between plate readings		
while time_counter < len(dates):															## creates a list with the difference between time points
	tdelta = dates[time_counter] - dates[time_counter-1]
	delta_time_points.append(tdelta)
	time_counter = time_counter+1

#### converts the difference in times into hours and minutes
def hours_minutes(td):																		## converts the delta timepoints into hours and minutes
	hours_out = []
	minutes_out = []
	for x in td:
		hours = x.seconds//3600
		minutes = (x.seconds//60)%60
		minutes_out.append(minutes)
		hours_out.append(hours)
	return hours_out, minutes_out

hours, minutes = hours_minutes(delta_time_points)

hour_percent = []
for x in minutes:																			## converts minutes into percent of hours
	temp = float(x)/60
	hour_percent.append(temp)

## combines hours and minutes
temp = [float(x)+y for x,y in zip(hours,hour_percent)]

## creates the timepoints from the difference between sampled times and and initial timepoint
x_axis = []
counter = 0
x_temp = []
initial_time = 16.72
x_axis.append(initial_time)
while counter < len(temp):
	x_temp = x_axis[counter]+temp[counter]
	x_axis.append(x_temp)
	counter = counter+1

well_value = [float(i) for i in well_value]

chunks=[well_value[x:x+96] for x in xrange(0, len(well_value), 96)]

counter=0
corrected_chunk = []
for x in chunks:
	corrected_chunk_temp = []
	for y in x:
		corrected_temp = (y-blank_average)*dilution_factor[counter]
		counter = counter+1
		corrected_chunk_temp.append(corrected_temp)
	corrected_chunk.append(corrected_chunk_temp)
	
Corrected_OD_location = out_dir + "/Corrected_OD.xlsx"
workbook = xlsxwriter.Workbook(Corrected_OD_location)
ws = workbook.add_worksheet()
ws.write(0,0, 'Time')
ws.write(0,1, 'Temperature')
#print wells_in_plate
ws.write_row('C1', wells_in_plate)
ws.write_column('A2', x_axis)
row_counter = 0
while row_counter < len(corrected_chunk):
	ws.write_row(row_counter+1,2,corrected_chunk[row_counter])
	row_counter = row_counter+1
workbook.close()
######################

workbook = xlsxwriter.Workbook(args.out_file)
worksheet = workbook.add_worksheet()
worksheet.write(0,0, 'strain')
worksheet.write(0,1, 'slope')
worksheet.write(0,2, 'standard error of the slope')
worksheet.write(0,3, 'correlation coefficient')
worksheet.write(0,4, 'hours used in analysis')




bar_chart = workbook.add_chart({'type': 'column'})
bar_chart.add_series({'name': 'Growth Rate', 'categories': '=Sheet1!$A$2:$A$11','values': '=Sheet1!$B$2:$B$11', 'y_error_bars': {'type': 'custom', 'plus_values': '=Sheet1!$C$2:$C$11', 'minus_values': '=Sheet1!$C$2:$C$11'}})
bar_chart.set_y_axis({'name': 'Growth rate'})
bar_chart.set_x_axis({'name': 'Strain'})
bar_chart.set_legend({'none': True})
bar_chart.set_style(11)
worksheet.insert_chart('G5', bar_chart)



counter2=0
row_counter = 0
while counter2 < (len(strain_name_list) - 1):
	row_counter = row_counter +1
	worksheet.write(row_counter,0, strain_name_list[counter2+1])
	worksheet.write(row_counter, 1, slope_list[counter2])
	worksheet.write(row_counter, 2, std_err_list[counter2])
	worksheet.write(row_counter, 3, r_value_list[counter2])
	worksheet.write(row_counter, 4, time_used[counter2])
	counter2 = counter2 + 1
	
	
counter3=0


while counter3 < (len(strain_name_list) - 1):
	worksheet2 = workbook.add_worksheet(strain_name_list[counter3 +1])
	worksheet2.insert_image('A2', 'Growth Rate for %s.png' % strain_name_list[counter3 + 1])	
	counter4=0
	x_var = []

	while counter4 < len(od_value_out[counter3]):
		counter5=1
		od_yvar = []
		for x in od_value_out[counter3]:
			worksheet2.write(0,13, 'corrected OD')
			worksheet2.write(counter5, 13, x)
			od_yvar.append(x)
			counter5=counter5+1
		counter5 = 1
	
		log_od_yvar = []
		for y in log_od_value_out[counter3]:
			worksheet2.write(0,14, 'log(corrected OD)')
			worksheet2.write(counter5, 14, y)
			log_od_yvar.append(y)
			counter5 = counter5+1
		counter5 = 1
		A600_yvar = []
		
		worksheet2.write(0,12, 'hour sampled')
		worksheet2.write(counter4+1,12, hour[counter4])
		x_var.append(hour[counter4])
		#print x_var
		#print ''
		counter4 = counter4+1

	plt.scatter(x_var,log_od_yvar)
	plt.title('%s' % strain_name_list[counter3+1])
	plt.xlabel('hours')
	plt.ylabel('log OD')
	plt.savefig('Log Growth Curve for %s' % strain_name_list[counter3 + 1])
	worksheet2.insert_image('A30', 'Log Growth Curve for %s.png' % strain_name_list[counter3 + 1])
	plt.close()
	
	plt.scatter(x_var,od_yvar)
	plt.title('%s' % strain_name_list[counter3+1])
	plt.xlabel('hours')
	plt.ylabel('Corrected OD')
	plt.savefig('Growth Curve for %s' % strain_name_list[counter3 + 1])
	worksheet2.insert_image('A60', 'Growth Curve for %s.png' % strain_name_list[counter3 + 1])
	plt.close()
	counter3 = counter3+1


plate_averages = plate_averages(well_value, well_name, input_array)


#print plate_averages
#print plate_averages[0]
#print plate_averages[0][0]
#print plate_averages[0][0][0]


info_count = 0
well_num = len(well_value)
worksheet3 = workbook.add_worksheet('all strains')
worksheet3.write(0,0, 'strain')
worksheet3.write(0,1, 'hour')
worksheet3.write(0,2, 'average od')


time_count1 = 0
row_counter = 0
while time_count1 < len(input_array[0]):
	strain_count1 = 0
	worksheet3.write(row_counter,6, 'Hour')
	worksheet3.write(row_counter,7, 'Strain')
	worksheet3.write(row_counter,8, 'Mean')
	worksheet3.write(row_counter,9, 'Std. Dev')
	worksheet3.write(row_counter,10, '% CV')
	row_counter = row_counter+1
	while strain_count1 < (len(input_array[2])-1):
		#print plate_averages[strain_count1][time_count1][3]
		worksheet3.write(row_counter, 6, plate_averages[strain_count1][time_count1][1])
		worksheet3.write(row_counter, 7, plate_averages[strain_count1][time_count1][0])
		worksheet3.write(row_counter, 8, plate_averages[strain_count1][time_count1][2])
		worksheet3.write(row_counter, 9, plate_averages[strain_count1][time_count1][3])
		CV_temp = float((plate_averages[strain_count1][time_count1][3])/(plate_averages[strain_count1][time_count1][2]))*100
		worksheet3.write(row_counter, 10, '%s %%' % CV_temp)
		row_counter = row_counter+1
		strain_count1 = strain_count1 + 1
	row_counter = row_counter+1
	time_count1 =time_count1 + 1
	
row_counter=1
strain_count = 0
y_value = []
x_value = []
while strain_count < len(plate_averages):
	time_count = 0
	x_value_temp = []
	y_value_temp= []
	while time_count < len(plate_averages[strain_count]):
		worksheet3.write(row_counter, 0, plate_averages[strain_count][time_count][0])
		worksheet3.write(row_counter, 1, plate_averages[strain_count][time_count][1])
		worksheet3.write(row_counter, 2, plate_averages[strain_count][time_count][2])
		x_value_temp.append(plate_averages[strain_count][time_count][1])
		y_value_temp.append(plate_averages[strain_count][time_count][2])
		row_counter = row_counter+1
		time_count = time_count+1
	y_value.append(y_value_temp)
	x_value.append(x_value_temp)
	strain_count = strain_count+1

strain_name= strain_name_list[1:]
marker = itertools.cycle(('.','o','v','^','<','>','1','2','3','4','8','s','p','*'))
for n in y_value:
	plt.plot(x_value[0],n, marker = marker.next(), markersize=7)
	plt.legend(strain_name, loc=4, fontsize = 10)
	plt.grid(True)

plt.title('All Strains')
plt.xlabel('hours')
plt.ylabel('OD')
plt.savefig('All Strains')
worksheet3.insert_image('N3', 'All Strains.png')
plt.close

workbook.close()

counter3=0
while counter3 < (len(strain_name_list) - 1):
	os.remove('Growth Rate for %s.png' % strain_name_list[counter3 + 1])
	os.remove('Log Growth Curve for %s.png' % strain_name_list[counter3 + 1])
	os.remove('Growth Curve for %s.png' % strain_name_list[counter3 + 1])
	counter3 = counter3 +1
os.remove('All Strains.png')
	